#update BYL
#update def getThruTempSite(BYL) filePath
#update getDataLOBYL(dLOBio5,dLO)
#update getCarrierBYL(schedBio5,metBio5,eoD)
#update getExcelBYL(SUMMARY INSERTION:,THRUPUT INSERTION:)
import openpyxl
import os
import csv
import xlrd
import re
import sys
sys.path.insert(0, '../Date')
import allDateFormat
from allDateFormat import getFullDate,getDate,getHour,getDateDash,getDashDate,getDateSlash
def getSumTempSite(x):
    if x == 'BLG':
        filePath = '../../template/thrusum/BLG/LAPORAN SUMMARY 15 Januari 2018 (2).xlsx'
        return filePath
def getThruTempSite(x):
    if x == 'PLP':
        filePath = '../../template/thrusum/PLP/laptru2.xlsx'
        return filePath
    elif x == 'BLG':
        filePath = '../../template/thrusum/BLG/LAPORAN THRUPUT 07 Desember 2017.xlsx'
        return filePath
    elif x == 'BYL':
	#filePath = '../../template/thrusum/BYL/Thruput 06 Desember 2017.xlsx'
        filePath = '../../template/thrusum/BYL/Laporan Thruput 23072018 .xlsx'
        return filePath
    elif x == 'KTP':
        filePath = '../../template/thrusum/KTP/12 Desember 2017 Laporan Thruput.xlsx'
        return filePath
    elif x == 'MDN':
        filePath = '../../template/thrusum/MDN/Laporan Thruput 12 Desember  2017.xlsx'
        return filePath
    elif x == 'PJG':
        filePath = '../../template/thrusum/PJG/laptru2.xlsx' #notYet
        return filePath
    elif x == 'PMB':
        filePath = '../../template/thrusum/PMB/laptru2.xlsx' #notYet
        return filePath
    elif x == 'SBY':
        filePath = '../../template/thrusum/SBY/Laporan Thruput 12 DESEMBER 2017.xlsx'
        return filePath
    elif x == 'TGR':
        filePath = '../../template/thrusum/TGR/Laporan Thruput 12-12-2017.xlsx'
        return filePath
    elif x == 'UJB':
        filePath = '../../template/thrusum/UJB/Laporan Thruput 12-12-2017.xlsx'
        return filePath
def getCsvMeter(path):
    exFile = open(path)
    exReader = csv.reader(exFile)
    csvLap = list(exReader)
    del csvLap[0]
    return csvLap
def getCsvCarr(path):
    xFile = open(path)
    xReader = csv.reader(xFile)
    eod = list(xReader)
    return eod
def getDataLO(path):
    xlFile = xlrd.open_workbook(path)
    dataLO = xlFile.sheet_by_index(0)
    return dataLO
def getByMeterPLP(csvFilePath):
    csvLap = getCsvMeter(csvFilePath)
    prodPre = []
    prodPx = []
    prodPxT = []
    prodDex = []
    prodSol = []
    prodFame = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if any([i[2] == 'PREMIUM',i[2] == 'PREMIUM02']):
            if all([i[0] != 302,i[0] != 502, i[0] != 602, i[0] != 601,i[0] != 802, i[0] != 1002, i[0] != 1102, i[0] != 1202]):
                prodPre.append(i)
        elif i[2] == 'SOLAR':
            if all([i[0] != 1301, i[0] != 1302, i[0] != 1303, i[0] != 1304]):
                prodSol.append(i)
        elif i[2] == 'FAME':
            if all([i[0] != 105, i[0] != 205,i[0] != 504, i[0] != 505,i[0] != 704, i[0] != 805,i[0] != 806,i[0] != 808,i[0] != 906,i[0] != 1105,i[0] != 1106,i[0] != 1107,i[0] != 1306,i[0] != 1307,i[0] != 1308,i[0] != 1309]):
                prodFame.append(i)
        elif i[2] == 'PERTAMAX':
            prodPx.append(i)
        elif i[2] == 'PERTAMAX TURBO':
            prodPxT.append(i)
        elif i[2] == 'PERTAMINA DEX':
            prodDex.append(i)
    #sort all
    prodPre.sort()
    prodPx.sort()
    prodPxT.sort()
    prodDex.sort()
    prodSol.sort()
    prodFame.sort()
    #final csv list
    akhirMeter = []
    for i in prodPre:
        akhirMeter.append(i[3])
    for i in prodPx:
        akhirMeter.append(i[3])
    for i in prodSol:
        akhirMeter.append(i[3])
    for i in prodFame:
        akhirMeter.append(i[3])
    for i in prodPxT:
        akhirMeter.append(i[3])
    for i in prodDex:
        akhirMeter.append(i[3])
    akhirMeter.insert(-1,None)
    return akhirMeter
def getByMeterBYL(csvFilePath):
    csvLap = getCsvMeter(csvFilePath)
    meterUnion = []
    #premium
    prodPre = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PREMIUM']):
                if any([i[0] == 101,i[0] == 104, i[0] == 201, i[0] == 204,i[0] == 301,i[0] == 302,i[0] == 306,i[0] == 401,i[0] == 404]):
                    prodPre.append(i)
    prodPre.sort()
    for i in prodPre:
        meterUnion.append(i[3])
    #solar
    prodSol = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'SOLAR']):
                if any([i[0] == 103,i[0] == 303, i[0] == 403]):
                    prodSol.append(i)
    prodSol.sort()
    for i in prodSol:
        meterUnion.append(i[3])
    #fame
    prodFame = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'FAME']):
                if any([i[0] == 105,i[0] == 305, i[0] == 405]):
                    prodFame.append(i)
    prodFame.sort()
    for i in prodFame:
        meterUnion.append(i[3])
    #pertamax
    prodPx = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PERTAMAX']):
                if any([i[0] == 102,i[0] == 202, i[0] == 203, i[0] == 304, i[0] == 402]):
                    prodPx.append(i)
    prodPx.sort()
    for i in prodPx:
        meterUnion.append(i[3])
    #adjust the column
    meterUnion.insert(12,None)
    meterUnion.insert(16,None)
    return meterUnion
def getByMeterKTP(csvFilePath):
    csvLap = getCsvMeter(csvFilePath)
    meterUnion = []
    #premium
    prodPre = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PREMIUM']):
                if any([i[0] == 102,i[0] == 105, i[0] == 201, i[0] == 202,i[0] == 301,i[0] == 302,i[0] == 306,i[0] == 501]):
                    prodPre.append(i)
    prodPre.sort()
    for i in prodPre:
        meterUnion.append(i[3])
    #solar
    prodSol = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'SOLAR']):
                if any([i[0] == 104,i[0] == 203, i[0] == 204,i[0] == 304,i[0] == 401,i[0] == 402,i[0] == 502]):
                    prodSol.append(i)
    prodSol.sort()
    for i in prodSol:
        meterUnion.append(i[3])
    #fame
    prodFame = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'FAME']):
                if any([i[0] == 106,i[0] == 205, i[0] == 206,i[0] == 305,i[0] == 403,i[0] == 404]):
                    prodFame.append(i)
    prodFame.sort()
    for i in prodFame:
        meterUnion.append(i[3])
    #pertamax
    prodPx = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PERTAMAX']):
                if any([i[0] == 103,i[0] == 303]):
                    prodPx.append(i)
    prodPx.sort()
    for i in prodPx:
        meterUnion.append(i[3])
    #pertaminaDex
    prodDex = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PERTAMINA DEX']):
                if any([i[0] == 405]):
                    prodDex.append(i)
    prodDex.sort()
    for i in prodDex:
        meterUnion.append(i[3])
    #adjust the column
    meterUnion.insert(12,None)
    meterUnion.insert(21,None)
    meterUnion.insert(24,None)
    meterUnion.insert(26,None)
    return meterUnion
def getByMeterMDN(csvFilePath):
    csvLap = getCsvMeter(csvFilePath)
    meterUnion = []
    #premium
    prodPre = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PREMIUM']):
                if any([i[0] == 101,i[0] == 201, i[0] == 301, i[0] == 401]):
                    prodPre.append(i)
    prodPre.sort()
    for i in prodPre:
        meterUnion.append(i[3])    
    #solar
    prodSol = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'SOLAR']):
                if any([i[0] == 103,i[0] == 203, i[0] == 303, i[0] == 403, i[0] == 504, i[0] == 801, i[0] == 802]):
                    prodSol.append(i)
    prodSol.sort()
    for i in prodSol:
        meterUnion.append(i[3])
    #prePl
    prodPrePl = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PREMIUM']):
                if any([i[0] == 104,i[0] == 207, i[0] == 304, i[0] == 405, i[0] == 503, i[0] == 505, i[0] == 703, i[0] == 704]):
                    prodPrePl.append(i)
    prodPrePl.sort()
    for i in prodPrePl:
        meterUnion.append(i[3])
    #fame
    prodFame = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'FAME']):
                if any([i[0] == 105,i[0] == 205, i[0] == 305, i[0] == 404, i[0] == 506]):
                    prodFame.append(i)
    prodFame.sort()
    for i in prodFame:
        meterUnion.append(i[3])
    #pertamax
    prodPx = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PERTAMAX']):
                if any([i[0] == 102,i[0] == 202, i[0] == 302, i[0] == 402, i[0] == 501, i[0] == 502, i[0] == 701, i[0] == 702]):
                    prodPx.append(i)
    prodPx.sort()
    for i in prodPx:
        meterUnion.append(i[3])
    #adjust the column
    meterUnion.insert(13,None)
    meterUnion.insert(15,None)
    meterUnion.insert(18,None)
    return meterUnion
def getByMeterSBY(csvFilePath):
    csvLap = getCsvMeter(csvFilePath)
    meterUnion = []
    #premium
    prodPre = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PREMIUM']):
                if any([i[0] == 201,i[0] == 301, i[0] == 401, i[0] == 501, i[0] == 602, i[0] == 801, i[0] == 901, i[0] == 1001]):
                    prodPre.append(i)
    prodPre.sort()
    for i in prodPre:
        meterUnion.append(i[3])    
    #prePl
    prodPrePl = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PREMIUM']):
                if any([i[0] == 204,i[0] == 205, i[0] == 308, i[0] == 406, i[0] == 407, i[0] == 505, i[0] == 608, i[0] == 805, i[0] == 907, i[0] == 1006]):
                    prodPrePl.append(i)
    prodPrePl.sort()
    for i in prodPrePl:
        meterUnion.append(i[3])
    #solar
    prodSol = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'SOLAR']):
                if any([i[0] == 303,i[0] == 503, i[0] == 603, i[0] == 701, i[0] == 804, i[0] == 904, i[0] == 1002]):
                    prodSol.append(i)
    prodSol.sort()
    for i in prodSol:
        meterUnion.append(i[3])

    #fame
    prodFame = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'FAME']):
                if any([i[0] == 305,i[0] == 506, i[0] == 607, i[0] == 703, i[0] == 808, i[0] == 909, i[0] == 1005]):
                    prodFame.append(i)
    prodFame.sort()
    for i in prodFame:
        meterUnion.append(i[3])
    #pertamax
    prodPx = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PERTAMAX']):
                if any([i[0] == 202,i[0] == 203, i[0] == 302, i[0] == 402, i[0] == 403, i[0] == 502, i[0] == 604, i[0] == 802, i[0] == 902, i[0] == 903, i[0] == 1003]):
                    prodPx.append(i)
    prodPx.sort()
    for i in prodPx:
        meterUnion.append(i[3])
    #pertamaxTurbo
    prodPxt = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PERTAMAX TURBO']):
                if any([i[0] == 601]):
                    prodPxt.append(i)
    prodPxt.sort()
    for i in prodPxt:
        meterUnion.append(i[3])
    #pertaminaDex
    prodDex = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PERTAMINA DEX']):
                if any([i[0] == 702,i[0] == 803]):
                    prodDex.append(i)
    prodDex.sort()
    for i in prodDex:
        meterUnion.append(i[3])
    
    #adjust the column
    meterUnion.insert(5,None)
    meterUnion.insert(19,0)
    meterUnion.insert(20,0)
    meterUnion.insert(21,0)
    meterUnion.insert(22,0)
    meterUnion.insert(24,None)
    meterUnion.insert(25,None)
    meterUnion.insert(26,None)
    meterUnion.insert(28,None)
    meterUnion.insert(30,None)
    meterUnion.insert(32,None)
    meterUnion.insert(34,None)
    meterUnion.insert(36,None)
    meterUnion.insert(38,None)
    meterUnion.insert(59,None)
    return meterUnion
def getByMeterTGR(csvFilePath):
    csvLap = getCsvMeter(csvFilePath)
    meterUnion = []
    #premium
    prodPre = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PREMIUM']):
                if any([i[0] == 101,i[0] == 102, i[0] == 206, i[0] == 208,i[0] == 301,i[0] == 304]):
                    prodPre.append(i)
    prodPre.sort()
    for i in prodPre:
        meterUnion.append(i[3])
    #solar
    prodSol = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'SOLAR']):
                if any([i[0] == 103,i[0] == 104, i[0] == 203, i[0] == 303, i[0] == 402]):
                    prodSol.append(i)
    prodSol.sort()
    for i in prodSol:
        meterUnion.append(i[3])
    #fame
    prodFame = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'FAME']):
                if any([i[0] == 107,i[0] == 108, i[0] == 207, i[0] == 305]):
                    prodFame.append(i)
    prodFame.sort()
    for i in prodFame:
        meterUnion.append(i[3])
    #pertamax
    prodPx = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PERTAMAX']):
                if any([i[0] == 202,i[0] == 204, i[0] == 302]):
                    prodPx.append(i)
    prodPx.sort()
    for i in prodPx:
        meterUnion.append(i[3])
    #MFO
    prodMfo = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'MFO180']):
                if any([i[0] == 401]):
                    prodMfo.append(i)
    prodMfo.sort()
    for i in prodMfo:
        meterUnion.append(i[3])
    #adjusting the none type to meterUnion
    meterUnion.insert(7,None)
    meterUnion.insert(9,None)
    meterUnion.insert(11,None)
    meterUnion.insert(21,None)
    return meterUnion
def getByMeterUJB(csvFilePath):
    csvLap = getCsvMeter(csvFilePath)
    meterUnion = []
    #premium
    prodPre = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PREMIUM']):
                if any([i[0] == 102,i[0] == 103, i[0] == 105, i[0] == 201,i[0] == 206,i[0] == 301,i[0] == 302,i[0] == 402,i[0] == 403,i[0] == 404]):
                    prodPre.append(i)
    prodPre.sort()
    for i in prodPre:
        meterUnion.append(i[3])
    #solar
    prodSol = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'SOLAR']):
                if any([i[0] == 104,i[0] == 202, i[0] == 303, i[0] == 304]):
                    prodSol.append(i)
    prodSol.sort()
    for i in prodSol:
        meterUnion.append(i[3])
    #fame
    prodFame = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'FAME']):
                if any([i[0] == 108,i[0] == 205, i[0] == 307, i[0] == 308]):
                    prodFame.append(i)
    prodFame.sort()
    for i in prodFame:
        meterUnion.append(i[3])
    #pertamax
    prodPx = []
    for i in csvLap:
        i[0] = int(i[0])
        i[3] = int(i[3])
        if([i[2] == 'PERTAMAX']):
                if any([i[0] == 101,i[0] == 203, i[0] == 401]):
                    prodPx.append(i)
    prodPx.sort()
    for i in prodPx:
        meterUnion.append(i[3])
    #adjust the column
    meterUnion.insert(11,None)
    meterUnion.insert(13,None)
    meterUnion.insert(17,None)
    meterUnion.insert(20,None)
    meterUnion.insert(22,None)
    return meterUnion
def getByMeterBLG(csvFilePath):
    csvLap = getCsvMeter(csvFilePath)
    meterUnionOp = []
    meterUnionCl = []
    #premium
    prodPreOp = []
    prodPreCl = []
    sortedProdPreOp = []
    sortedProdPreCl = []
    for i in csvLap:
        i[5] = i[5].replace(",","")
        i[8] = i[8].replace(",","")
        if any([i[2] == '202',i[2] == '205',i[2] == '301', i[2] == '304', i[2] == '401',i[2] == '404']):
            prodPreOp.append(i)
            prodPreCl.append(i)
    # sort the opening
    prodPreOp.sort(key = lambda x: x[2])
    for i in prodPreOp:
        sortedProdPreOp.append(int(i[5]))
    # sort the closing
    prodPreCl.sort(key = lambda x: x[2])
    for i in prodPreCl:
        sortedProdPreCl.append(int(i[8]))
    for i in sortedProdPreOp:
        meterUnionOp.append(i)
    for i in sortedProdPreCl:
        meterUnionCl.append(i)
    #solar
    prodSolOp = []
    prodSolCl = []
    for i in csvLap:
        i[5] = i[5].replace(",","")
        i[8] = i[8].replace(",","")
        if any([i[2] == '204',i[2] == '303', i[2] == '403']):
            prodSolOp.append(int(i[5]))
            prodSolCl.append(int(i[8]))
    for i in prodSolOp:
        meterUnionOp.append(i)
    for i in prodSolCl:
        meterUnionCl.append(i)
    #fame
    prodFameCl = []
    prodFameOp = []
    for i in csvLap:
        i[5] = i[5].replace(",","")
        i[8] = i[8].replace(",","")
        if any([i[2] == '206',i[2] == '305', i[2] == '405']):
            prodFameOp.append(int(i[5]))
            prodFameCl.append(int(i[8]))
    for i in prodFameOp:
        meterUnionOp.append(i)
    for i in prodFameCl:
        meterUnionCl.append(i)
    #pertamax
    prodPxCl = []
    prodPxOp = []
    for i in csvLap:
        i[5] = i[5].replace(",","")
        i[8] = i[8].replace(",","")
        if any([i[2] == '203',i[2] == '302', i[2] == '402']):
            prodPxOp.append(int(i[5]))
            prodPxCl.append(int(i[8]))
    for i in prodPxOp:
        meterUnionOp.append(i)
    for i in prodPxCl:
        meterUnionCl.append(i)
    #pertamaxTurbo
    prodPxtCl = []
    prodPxtOp = []
    for i in csvLap:
        i[5] = i[5].replace(",","")
        i[8] = i[8].replace(",","")
        if any([i[2] == '201']):
            prodPxtOp.append(int(i[5]))
            prodPxtCl.append(int(i[8]))
    for i in prodPxtOp:
        meterUnionOp.append(i)
    for i in prodPxtCl:
        meterUnionCl.append(i)
    #Avtur
    prodAvtCl = []
    prodAvtOp = []
    for i in csvLap:
        i[5] = i[5].replace(",","")
        i[8] = i[8].replace(",","")
        if any([i[2] == '101',i[2] == '102']):
            prodAvtOp.append(int(i[5]))
            prodAvtCl.append(int(i[8]))
    for i in prodAvtOp:
        meterUnionOp.append(i)
    for i in prodAvtCl:
        meterUnionCl.append(i)
    #adjust the column
    meterUnionOp.insert(7,None)
    meterUnionOp.insert(8,None)
    meterUnionOp.insert(9,None)
    meterUnionOp.insert(11,None)
    meterUnionOp.insert(13,None)
    meterUnionOp.insert(14,None)
    meterUnionOp.insert(19,None)
    meterUnionOp.insert(21,None)
    meterUnionOp.insert(24,None)
    meterUnionOp.insert(25,None)
    meterUnionOp.insert(26,None)
    meterUnionOp.insert(27,None)
    meterUnionOp.insert(28,None)
    meterUnionCl.insert(7,None)
    meterUnionCl.insert(8,None)
    meterUnionCl.insert(9,None)
    meterUnionCl.insert(11,None)
    meterUnionCl.insert(13,None)
    meterUnionCl.insert(14,None)
    meterUnionCl.insert(19,None)
    meterUnionCl.insert(21,None)
    meterUnionCl.insert(24,None)
    meterUnionCl.insert(25,None)
    meterUnionCl.insert(26,None)
    meterUnionCl.insert(27,None)
    meterUnionCl.insert(28,None)
    return meterUnionOp,meterUnionCl
def getDataLOPLP(dLOFilePath):
    dataLO = getDataLO(dLOFilePath)
    listPre = []
    listPx = []
    listPxT = []
    listDexl = []
    listPDex = []
    listPlt = []
    listSol = []
    listBio = []
    #product summary by data LO
    for i in range(3, dataLO.nrows):
        if dataLO.cell_value(i,12) == 'PREMIUM':
            listPre.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12) == 'BIOSOLAR':
            listBio.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12) == 'SOLAR':
            listSol.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12) == 'DEXLITE':
            listDexl.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12) == 'PERTAMINA-DEX':
            listPDex.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12) == 'PERTAMAX':
            listPx.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12) == 'PERTAMAX-TURBO':
            listPxT.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12) == 'PERTALITE':
            listPlt.append(dataLO.cell_value(i,14))
    dLO = {'PREMIUM':int(sum(listPre)),
          'SOLAR':int(sum(listSol)),
          'BIOSOLAR':int(sum(listBio)),
          'PERTAMAX':int(sum(listPx)),
          'PERTAMAX TURBO':int(sum(listPxT)),
          'PERTALITE':int(sum(listPlt)),
          'PERTAMINA DEX':int(sum(listPDex)),
          'DEXLITE':int(sum(listDexl))}
    return dLO
def getDataLOBLG(dLOFilePath):
    dataLO = getDataLO(dLOFilePath)
    dLOPre = []
    dLOSol = []
    dLOBio = []
    dLOBioP= []
    dLOPx = []
    dLOPxt= []
    dLOPl = []
    dLOAvt = []
    #BYLO PROCESSING
    for i in range(3, dataLO.nrows):
        if dataLO.cell_value(i,12)=='PREMIUM':
            dLOPre.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='SOLAR':
            dLOSol.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='BIOSOLAR':
            dLOBio.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='BIOSOLAR PEMBANGKIT LISTRIK':
            dLOBioP.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='PERTAMAX':
            dLOPx.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='PERTAMAX-TURBO':
            dLOPxt.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='PERTALITE':
            dLOPl.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='AVTUR':
            dLOAvt.append(dataLO.cell_value(i,14))
    dLO = {'PREMIUM':int(sum(dLOPre)),
          'SOLAR':int(sum(dLOSol)),
          'BIOSOLAR':int(sum(dLOBio)),
          'BIOSOLAR PEMBANGKIT':int(sum(dLOBioP)),
          'PERTAMAX':int(sum(dLOPx)),
          'PERTAMAX TURBO':int(sum(dLOPxt)),
          'PERTALITE':int(sum(dLOPl)),
          'AVTUR':int(sum(dLOAvt))}
    return dLO
def getDataLOBYL(dLOFilePath):
    dataLO = getDataLO(dLOFilePath)
    dLOPre = []
    dLOSol = []
    dLOBio = []
    dLOPx = []
    dLOPl = []
    dLOBio5 = []
    for i in range(3, dataLO.nrows):
        if dataLO.cell_value(i,12)=='PREMIUM':
            dLOPre.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='SOLAR':
            dLOSol.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='BIOSOLAR':
            dLOBio.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='PERTAMAX':
            dLOPx.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='PERTALITE':
            dLOPl.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,12)=='BIOSOLAR B5':
                dLOBio5.append(dataLO.cell_value(i,14))
    dLO = {'PREMIUM':int(sum(dLOPre)),
          'SOLAR':int(sum(dLOSol)),
          'BIOSOLAR':int(sum(dLOBio)),
          'PERTAMAX':int(sum(dLOPx)),
          'PERTALITE':int(sum(dLOPl)),
		  'BIOSOLAR B5':int(sum(dLOBio5))}
    return dLO
def getDataLOKTP(dLOFilePath):
    dataLO = getDataLO(dLOFilePath)
    dLOBio = []
    dLODexl = []
    dLOPx = []
    dLODex = []
    dLOPl = []
    dLOBioP = []
    dLOSolP = []
    for i in range(3, dataLO.nrows):
        if dataLO.cell_value(i,13)=='BIOSOLAR':
            dLOBio.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='DEXLITE':
            dLODexl.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='PERTAMAX':
            dLOPx.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='PERTAMINA-DEX':
            dLODex.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='PERTALITE':
            dLOPl.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='BIOSOLAR PEMBANGKIT LISTRIK':
            dLOBioP.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='SOLAR PEMBANGKIT LISTRIK':
            dLOSolP.append(dataLO.cell_value(i,14))
    dLO = {'DEXLITE': int(sum(dLODexl)),
           'PERTAMINA-DEX': int(sum(dLODex)),
           'BIOSOLAR': int(sum(dLOBio)),
           'PERTAMAX': int(sum(dLOPx)),
           'PERTALITE': int(sum(dLOPl)),
           'BIOSOLAR-PL':int(sum(dLOBioP)),
           'SOLAR-PL':int(sum(dLOSolP))}
    return dLO
def getDataLOMDN(dLOFilePath):
    dataLO = getDataLO(dLOFilePath)
    dLOPre = []
    dLOSol = []
    dLOBio = []
    dLOPx = []
    dLOPl = []
    dLODexL = []
    dLOSolRTW = []
    dLOPlRTW = []
    for i in range(3, dataLO.nrows):
        if dataLO.cell_value(i,13)=='PREMIUM':
            dLOPre.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='SOLAR':
            dLOSol.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='BIOSOLAR':
            dLOBio.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='PERTAMAX':
            dLOPx.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='PERTALITE':
            dLOPl.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='DEXLITE':
            dLODexL.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='SOLAR RTW':
            dLOSolRTW.append(dataLO.cell_value(i,14))
        elif dataLO.cell_value(i,13)=='PERTALITE RTW':
            dLOPlRTW.append(dataLO.cell_value(i,14))
    dLO = {'PREMIUM':int(sum(dLOPre)),
          'SOLAR':int(sum(dLOSol)),
          'BIOSOLAR':int(sum(dLOBio)),
          'PERTAMAX':int(sum(dLOPx)),
          'PERTALITE':int(sum(dLOPl)),
          'DEXLITE':int(sum(dLODexl)),
          'SOLAR RTW':int(sum(dLOSolRTW)),
          'PERTALITE RTW':int(sum(dLOPlRTW))}
    return dLO
def getDataLOSBY(dLOFilePath):
    dataLO = getDataLO(dLOFilePath)
    dLOPre = []
    dLOSol = []
    dLOBio = []
    dLOBioP= []
    dLOPx = []
    dLOPxt= []
    dLOPl = []
    dLODex= []
    dLODexl=[]
    for i in range(3, dataLO.nrows):
        if dataLO.cell_value(i,17)=='PREMIUM':
            dLOPre.append(dataLO.cell_value(i,18))
        elif dataLO.cell_value(i,17)=='SOLAR':
            dLOSol.append(dataLO.cell_value(i,18))
        elif dataLO.cell_value(i,17)=='BIOSOLAR':
            dLOBio.append(dataLO.cell_value(i,18))
        elif dataLO.cell_value(i,17)=='BIOSOLAR PEMBANGKIT LISTRIK':
            dLOBioP.append(dataLO.cell_value(i,18))
        elif dataLO.cell_value(i,17)=='PERTAMAX':
            dLOPx.append(dataLO.cell_value(i,18))
        elif dataLO.cell_value(i,17)=='PERTAMAX-TURBO':
            dLOPxt.append(dataLO.cell_value(i,18))
        elif dataLO.cell_value(i,17)=='PERTALITE':
            dLOPl.append(dataLO.cell_value(i,18))
        elif dataLO.cell_value(i,17)=='PERTAMINA-DEX':
            dLODex.append(dataLO.cell_value(i,18))
        elif dataLO.cell_value(i,17)=='DEXLITE':
            dLODexl.append(dataLO.cell_value(i,18))
    dLO = {'PREMIUM':int(sum(dLOPre)),
          'SOLAR':int(sum(dLOSol)),
          'BIOSOLAR':int(sum(dLOBio)),
          'BIOSOLAR PEMBANGKIT':int(sum(dLOBioP)),
          'PERTAMAX':int(sum(dLOPx)),
          'PERTAMAX TURBO':int(sum(dLOPxt)),
          'PERTALITE':int(sum(dLOPl)),
          'PERTAMINA DEX':int(sum(dLODex)),
          'DEXLITE':int(sum(dLODexl))}
    return dLO
def getDataLOTGR(dLOFilePath):
    dataLO = getDataLO(dLOFilePath)
    dLOPre = []
    dLOSol = []
    dLOBio = []
    dLOBioP= []
    dLOPx = []
    dLOMFO= []
    dLOPl = []
    for i in range(3, dataLO.nrows):
        if dataLO.cell_value(i,13)=='PREMIUM':
            dLOPre.append(dataLO.cell_value(i,15))
        elif dataLO.cell_value(i,13)=='SOLAR':
            dLOSol.append(dataLO.cell_value(i,15))
        elif dataLO.cell_value(i,13)=='BIOSOLAR':
            dLOBio.append(dataLO.cell_value(i,15))
        elif dataLO.cell_value(i,13)=='BIOSOLAR PEMBANGKIT LISTRIK':
            dLOBioP.append(dataLO.cell_value(i,15))
        elif dataLO.cell_value(i,13)=='PERTAMAX':
            dLOPx.append(dataLO.cell_value(i,15))
        elif dataLO.cell_value(i,13)=='MFO':
            dLOMFO.append(dataLO.cell_value(i,15))
        elif dataLO.cell_value(i,13)=='PERTALITE':
            dLOPl.append(dataLO.cell_value(i,15))
    dLO = {'PREMIUM':int(sum(dLOPre)),
          'SOLAR':int(sum(dLOSol)),
          'BIOSOLAR':int(sum(dLOBio)),
          'BIOSOLAR PEMBANGKIT':int(sum(dLOBioP)),
          'PERTAMAX':int(sum(dLOPx)),
          'MFO':int(sum(dLOMFO)),
          'PERTALITE':int(sum(dLOPl)),}
    return dLO
def getDataLOUJB(dLOFilePath):
    dataLO = getDataLO(dLOFilePath)
    dLOPre = []
    dLOSol = []
    dLOBio = []
    dLOPx = []
    dLOPl = []
    #BYLO PROCESSING
    for i in range(3, dataLO.nrows):
        if dataLO.cell_value(i,12)=='PREMIUM':
            dLOPre.append(dataLO.cell_value(i,13))
        elif dataLO.cell_value(i,12)=='SOLAR':
            dLOSol.append(dataLO.cell_value(i,13))
        elif dataLO.cell_value(i,12)=='BIOSOLAR':
            dLOBio.append(dataLO.cell_value(i,13))
        elif dataLO.cell_value(i,12)=='PERTAMAX':
            dLOPx.append(dataLO.cell_value(i,13))
        elif dataLO.cell_value(i,12)=='PERTALITE':
            dLOPl.append(dataLO.cell_value(i,13))
    dLO = {'PREMIUM':int(sum(dLOPre)),
           'SOLAR':int(sum(dLOSol)),
           'BIOSOLAR':int(sum(dLOBio)),
           'PERTAMAX':int(sum(dLOPx)),
           'PERTALITE':int(sum(dLOPl))}
    return dLO
def getCarrierPLP(csvCarrier):
    eod = getCsvCarr(csvCarrier)
    eodPre = []
    eodPx = []
    eodPl = []
    eodPxT = []
    eodDexl = []
    eodDex = []
    eodSol = []
    eodBio = []
    metSol = []
    metDexlite = []
    metPl = []
    metPxT = []
    metDex = []
    metPre = []
    metBio = []
    metPx = []
    for i in eod:
        i[9] = i[9].replace(",","")
        i[10] = i[10].replace(",","")
        if i[3] == 'PREMIUM':
            eodPre.append(int(i[9]))
            metPre.append(int(i[10])) 
        elif i[3] == 'PERTAMAX':
            eodPx.append(int(i[9]))
            metPx.append(int(i[9]))
        elif i[3] == 'PERTALITE':
            eodPl.append(int(i[9]))
            metPl.append(int(i[10]))
        elif i[3] == 'PERTAMAX TURBO':
            eodPxT.append(int(i[9]))
            metPxT.append(int(i[10]))
        elif i[3] == 'DEXLITE':
            eodDexl.append(int(i[9]))
            metDexlite.append(int(i[10]))
        elif i[3] == 'DEX':
            eodDex.append(int(i[9]))
            metDex.append(int(i[10]))
        elif i[3] == 'SOLAR':
            eodSol.append(int(i[9]))
            metSol.append(int(i[10]))
        elif i[3] == 'BIOSOLAR':
            eodBio.append(int(i[9]))
            metBio.append(int(i[10]))
    eoD = {'eodPREMIUM':sum(eodPre),
           'metPREMIUM':sum(metPre),
           'eodPERTAMAX':sum(eodPx),
           'metPERTAMAX':sum(metPx),
           'eodPERTALITE':sum(eodPl),
           'metPERTALITE':sum(metPl),
           'eodPERTAMAX-TURBO':sum(eodPxT),
           'metPERTAMAX-TURBO':sum(metPxT),
           'eodDEXLITE':sum(eodDexl),
           'metDEXLITE':sum(metDexlite),
           'eodDEX':sum(eodDex),
           'metDEX':sum(metDex),
           'eodSOLAR':sum(eodSol),
           'metSOLAR':sum(metSol),
           'eodBIOSOLAR':sum(eodBio),
           'metBIOSOLAR':sum(metBio)}
    return eoD
def getCarrierBLG(csvCarrier):
    eod = getCsvCarr(csvCarrier)
    schedPre = []
    metPre = []
    schedSol = []
    metSol = []
    schedPx = []
    metPx = []
    schedBio = []
    metBio = []
    schedBioP = []
    metBioP = []
    schedPxT = []
    metPxT = []
    schedPl = []
    metPl = []
    schedAvt = []
    metAvt = []
    for i in eod:
        i[9] = i[9].replace(",","")
        i[10]= i[10].replace(",","")
        if i[3] == 'PREMIUM':
            schedPre.append(int(i[9]))
            metPre.append(int(i[10]))
        elif i[3]=='SOLAR':
            schedSol.append(int(i[9]))
            metSol.append(int(i[10]))
        elif i[3]=='PERTAMAX':
            schedPx.append(int(i[9]))
            metPx.append(int(i[10]))
        elif i[3]=='BIOSOLAR':
            schedBio.append(int(i[9]))
            metBio.append(int(i[10]))
        elif i[3]=='BIOSOLAR PL':
            schedBioP.append(int(i[9]))
            metBioP.append(int(i[10]))
        elif i[3]=='PERTAMAX TURBO':
            schedPxT.append(int(i[9]))
            metPxT.append(int(i[10]))
        elif i[3]=='PERTALITE':
            schedPl.append(int(i[9]))
            metPl.append(int(i[10]))
        elif i[3]=='JET-A1':
            schedAvt.append(int(i[9]))
            metAvt.append(int(i[10]))
    eoD = {'eodPREMIUM':sum(schedPre),
           'metPREMIUM':sum(metPre),
           'eodSOLAR':sum(schedSol),
           'metSOLAR':sum(metSol),
           'eodBIOSOLAR':sum(schedBio),
           'metBIOSOLAR':sum(metBio),
           'eodBIOSOLAR-PL':sum(schedBioP),
           'metBIOSOLAR-PL':sum(metBioP),
           'eodPERTAMAX':sum(schedPx),
           'metPERTAMAX':sum(metPx),
           'eodPERTALITE':sum(schedPl),
           'metPERTALITE':sum(metPl),
           'eodPERTAMAX-TURBO':sum(schedPxT),
           'metPERTAMAX-TURBO':sum(metPxT),
           'eodAVTUR':sum(schedAvt),
           'metAVTUR':sum(metAvt)}
    return eoD
def getCarrierBYL(csvCarrier):
    eod = getCsvCarr(csvCarrier)
    schedPre = []
    metPre = []
    schedSol = []
    metSol = []
    schedPx = []
    metPx = []
    schedBio = []
    metBio = []
    schedPl = []
    metPl = []
    schedBio5 = []
    metBio5 = []
    for i in eod:
        i[9] = i[9].replace(",","")
        i[10]= i[10].replace(",","")
        if i[3] == 'PREMIUM':
            schedPre.append(int(i[9]))
            metPre.append(int(i[10]))
        elif i[3]=='SOLAR':
            schedSol.append(int(i[9]))
            metSol.append(int(i[10]))
        elif i[3]=='PERTAMAX':
            schedPx.append(int(i[9]))
            metPx.append(int(i[10]))
        elif i[3]=='BIOSOLAR':
            schedBio.append(int(i[9]))
            metBio.append(int(i[10]))
        elif i[3]=='PERTALITE':
            schedPl.append(int(i[9]))
            metPl.append(int(i[10]))
        elif i[3]=='BIOSOLAR B5':
            schedBio5.append(int(i[9]))
            metBio5.append(int(i[10]))
    eoD = {'eodPREMIUM':sum(schedPre),
           'metPREMIUM':sum(metPre),
           'eodSOLAR':sum(schedSol),
           'metSOLAR':sum(metSol),
           'eodBIOSOLAR':sum(schedBio),
           'metBIOSOLAR':sum(metBio),
           'eodPERTAMAX':sum(schedPx),
           'metPERTAMAX':sum(metPx),
           'eodPERTALITE':sum(schedPl),
           'metPERTALITE':sum(metPl),
           'eodBio5':sum(schedBio5),
           'metBio5':sum(metBio5)}
    return eoD
def getCarrierKTP(csvCarrier):
    eod = getCsvCarr(csvCarrier)
    schedPre = []
    metPre = []
    schedSol = []
    metSol = []
    schedPx = []
    metPx = []
    schedBio = []
    metBio = []
    schedDexl = []
    metDexl = []
    schedDex = []
    metDex = []
    schedPl = []
    metPl = []
    schedBioP = []
    metBioP = []
    schedSolP = []
    metSolP = []
    schedPreRTW = []
    metPreRTW = []
    schedSolRTW = []
    metSolRTW = []
    schedPreTK = []
    metPreTK = []
    schedSolTK = []
    metSolTK = []
    RTW = re.compile('RTW.*')
    TKG = re.compile('TKG.*')
    for i in eod:
        i[9] = i[9].replace(",","")
        i[10]= i[10].replace(",","")
        if i[3]=='PERTAMAX':
            schedPx.append(int(i[9]))
            metPx.append(int(i[10]))
        elif i[3]=='BIOSOLAR':
            schedBio.append(int(i[9]))
            metBio.append(int(i[10]))
        elif i[3]=='DEXLITE':
            schedDexl.append(int(i[9]))
            metDexl.append(int(i[10]))
        elif i[3]=='PERTAMINA DEX':
            schedDex.append(int(i[9]))
            metDex.append(int(i[10]))
        elif i[3]=='PERTALITE':
            schedPl.append(int(i[9]))
            metPl.append(int(i[10]))
        elif i[3]=='BIOSOLAR PL':
            schedBioP.append(int(i[9]))
            metBioP.append(int(i[10]))
        elif i[3]=='SOLAR PL':
            schedSolP.append(int(i[9]))
            metSolP.append(int(i[10]))
    for i in eod:
        if RTW.match(i[8]):
            if i[3] == 'PREMIUM':
                schedPreRTW.append(int(i[9]))
                metPreRTW.append(int(i[10]))
            elif i[3] == 'SOLAR':
                schedSolRTW.append(int(i[9]))
                metSolRTW.append(int(i[10]))
        elif TKG.match(i[8]):
            if i[3] == 'PREMIUM':
                schedPreTK.append(int(i[9]))
                metPreTK.append(int(i[9]))
            elif i[3] == 'SOLAR':
                schedSolTK.append(int(i[9]))
                metSolTK.append(int(i[10]))
        else:
            if i[3] =='PREMIUM':
                schedPre.append(int(i[9]))
                metPre.append(int(i[10]))
            elif i[3] =='SOLAR':
                schedSol.append(int(i[9]))
                metSol.append(int(i[10]))
    schedPreRtwTkg = sum(schedPreRTW) + sum(schedPreTK)
    schedSolRtwTkg = sum(schedSolRTW) + sum(schedSolTK)
    metPreRtwTkg = sum(metPreRTW) + sum(metPreTK)
    metSolRtwTkg = sum(metSolRTW) + sum(metSolTK)
    eoD = {'eodPREMIUM':sum(schedPre),
           'metPREMIUM':sum(metPre),
           'eodPREMIUM-RTW':sum(schedPreRTW),
           'metPREMIUM-RTW':sum(metPreRTW),
           'eodPREMIUM-TKG':sum(schedPreTK),
           'metPREMIUM-TKG':sum(metPreTK),
           'eodPREMIUM-RTW-TKG':schedPreRtwTkg,
           'metPREMIUM-RTW-TKG':metPreRtwTkg,
           'eodSOLAR':sum(schedSol),
           'metSOLAR':sum(metSol),
           'eodSOLAR-PL':sum(schedSolP),
           'metSOLAR-PL':sum(metSolP),
           'eodSOLAR-RTW':sum(schedSolRTW),
           'metSOLAR-RTW':sum(metSolRTW),
           'eodSOLAR-TKG':sum(schedSolTK),
           'metSOLAR-TKG':sum(metSolTK),
           'eodSOLAR-RTW-TKG':schedSolRtwTkg,
           'metSOLAR-RTW-TKG':metSolRtwTkg,
           'eodBIOSOLAR':sum(schedBio),
           'metBIOSOLAR':sum(metBio),
           'eodBIOSOLAR-PL':sum(schedBioP),
           'metBIOSOLAR-PL':sum(metBioP),
           'eodPERTAMAX':sum(schedPx),
           'metPERTAMAX':sum(metPx),
           'eodPERTALITE':sum(schedPl),
           'metPERTALITE':sum(metPl),
           'eodDEXLITE':sum(schedDexl),
           'metDEXLITE':sum(metDexl),
           'eodDEX':sum(schedDex),
           'metDEX':sum(metDex)}
    return eoD
def getCarrierMDN(csvCarrier):
    eod = getCsvCarr(csvCarrier)
    schedPre = []
    metPre = []
    schedSolRTW = []
    metSolRTW = []
    schedSol = []
    metSol = []
    schedPx = []
    metPx = []
    schedPlRTW = []
    metPlRTW = []
    schedPl = []
    metPl = []
    schedDexl = []
    metDexl = []
    for i in eod:
        i[9] = i[9].replace(",","")
        i[10]= i[10].replace(",","")
        if i[3] == 'PREMIUM':
            schedPre.append(int(i[9]))
            metPre.append(int(i[10]))
        elif i[3]=='SOLAR':
            if i[0]=='RTW':
                schedSolRTW.append(int(i[9]))
                metSolRTW.append(int(i[10]))
            else:
                schedSol.append(int(i[9]))
                metSol.append(int(i[10]))
        elif i[3]=='PERTAMAX':
            schedPx.append(int(i[9]))
            metPx.append(int(i[10]))
        elif i[3]=='BIOSOLAR':
            schedBio.append(int(i[9]))
            metBio.append(int(i[10]))
        elif i[3]=='PERTALITE':
            if i[0]=='RTW':
                schedPlRTW.append(int(i[9]))
                metPlRTW.append(int(i[10]))
            else:
                schedPl.append(int(i[9]))
                metPl.append(int(i[10]))
        elif i[3]=='DEXLITE':
            schedDexl.append(int(i[9]))
            metDexl.append(int(i[10]))
    eoD = {'eodPREMIUM':sum(schedPre),
           'metPREMIUM':sum(metPre),
           'eodSOLAR':sum(schedSol),
           'metSOLAR':sum(metSol),
           'eodSOLAR-RTW':sum(schedSolRTW),
           'metSOLAR-RTW':sum(metSolRTW),
           'eodBIOSOLAR':sum(schedBio),
           'metBIOSOLAR':sum(metBio),
           'eodPERTAMAX':sum(schedPx),
           'metPERTAMAX':sum(metPx),
           'eodPERTALITE':sum(schedPl),
           'metPERTALITE':sum(metPl),
           'eodPERTALITE-RTW':sum(schedPlRTW),
           'metPERTALITE-RTW':sum(metPlRTW),
           'eodDEXLITE':sum(schedDexl),
           'metDEXLITE':sum(metDexl)}
    return eoD
def getCarrierSBY(csvCarrier):
    eod = getCsvCarr(csvCarrier)
    schedPre = []
    schedSol = []
    schedPx = []
    schedBio = []
    schedBioP= []
    schedPxT = []
    schedPl = []
    schedDex = []
    schedDexl = []
    metPre = []
    metSol = []
    metPx = []
    metBio = []
    metBioP = []
    metPxT = []
    metPl = []
    metDex = []
    metDexl = []
    for i in eod:
        i[9] = i[9].replace(",","")
        i[10]= i[10].replace(",","")
        if i[3] == 'PREMIUM':
            schedPre.append(int(i[9]))
            metPre.append(int(i[10]))
        elif i[3]=='SOLAR':
            schedSol.append(int(i[9]))
            metSol.append(int(i[10]))
        elif i[3]=='PERTAMAX':
            schedPx.append(int(i[9]))
            metPx.append(int(i[10]))
        elif i[3]=='BIOSOLAR':
            schedBio.append(int(i[9]))
            metBio.append(int(i[10]))
        elif i[3]=='BIOSOLAR PL':
            schedBioP.append(int(i[9]))
            metBioP.append(int(i[10]))
        elif i[3]=='PERTAMAX TURBO':
            schedPxT.append(int(i[9]))
            metPxT.append(int(i[10]))
        elif i[3]=='PERTALITE':
            schedPl.append(int(i[9]))
            metPl.append(int(i[10]))
        elif i[3]=='PERTAMINA DEX':
            schedDex.append(int(i[9]))
            metDex.append(int(i[9]))
        elif i[3]=='DEXLITE':
            schedDexl.append(int(i[9]))
            metDexl.append(int(i[9]))
    eoD = {'eodPREMIUM':sum(schedPre),
           'metPREMIUM':sum(metPre),
           'eodSOLAR':sum(schedSol),
           'metSOLAR':sum(metSol),
           'eodBIOSOLAR':sum(schedBio),
           'metBIOSOLAR':sum(metBio),
           'eodBIOSOLAR-PL':sum(schedBioP),
           'metBIOSOLAR-PL':sum(metBioP),
           'eodPERTAMAX':sum(schedPx),
           'metPERTAMAX':sum(metPx),
           'eodPERTAMAX-TURBO':sum(schedPxT),
           'metPERTAMAX-TURBO':sum(metPxT),
           'eodPERTALITE':sum(schedPl),
           'metPERTALITE':sum(metPl),
           'eodDEXLITE':sum(schedDexl),
           'metDEXLITE':sum(metDexl),
           'eodDEX':sum(schedDex),
           'metDEX':sum(metDex)}
    return eoD
def getCarrierTGR(csvCarrier):
    eod = getCsvCarr(csvCarrier)
    schedPre = []
    schedSol = []
    schedPx = []
    schedBio = []
    schedPl = []
    schedBioP = []
    schedMFO = []
    metPre = []
    metSol = []
    metPx = []
    metBio = []
    metPl = []
    metBioP = []
    metMFO = []
    for i in eod:
        i[9] = i[9].replace(",","")
        i[10]= i[10].replace(",","")
        if i[3] == 'PREMIUM':
            schedPre.append(int(i[9]))
            metPre.append(int(i[10]))
        elif i[3]=='SOLAR':
            schedSol.append(int(i[9]))
            metSol.append(int(i[10]))
        elif i[3]=='PERTAMAX':
            schedPx.append(int(i[9]))
            metPx.append(int(i[10]))
        elif i[3]=='BIOSOLAR':
            schedBio.append(int(i[9]))
            metBio.append(int(i[10]))
        elif i[2]=='A040900028':
            schedBioP.append(int(i[9]))
            metBioP.append(int(i[10]))
        elif i[3]=='MFO180':
            schedMFO.append(int(i[9]))
            metMFO.append(int(i[10]))
        elif i[3]=='PERTALITE':
            schedPl.append(int(i[9]))
            metPl.append(int(i[10]))
    eoD = {'eodPREMIUM':sum(schedPre),
           'metPREMIUM':sum(metPre),
           'eodSOLAR':sum(schedSol),
           'metSOLAR':sum(metSol),
           'eodBIOSOLAR':sum(schedBio),
           'metBIOSOLAR':sum(metBio),
           'eodBIOSOLAR-PL':sum(schedBioP),
           'metBIOSOLAR-PL':sum(metBioP),
           'eodPERTAMAX':sum(schedPx),
           'metPERTAMAX':sum(metPx),
           'eodMFO':sum(schedMFO),
           'metMFO':sum(metMFO),
           'eodPERTALITE':sum(schedPl),
           'metPERTALITE':sum(metPl)}
    return eoD
def getCarrierUJB(csvCarrier):
    eod = getCsvCarr(csvCarrier)
    schedPre = []
    schedSol = []
    schedPx = []
    schedBio = []
    schedPl = []
    metPre = []
    metSol = []
    metPx = []
    metBio = []
    metPl = []
    for i in eod:
        i[9] = i[9].replace(",","")
        i[10]= i[10].replace(",","")
        if i[3] == 'PREMIUM':
            schedPre.append(int(i[9]))
            metPre.append(int(i[10]))
        elif i[3]=='SOLAR':
            schedSol.append(int(i[9]))
            metSol.append(int(i[10]))
        elif i[3]=='PERTAMAX':
            schedPx.append(int(i[9]))
            metPx.append(int(i[10]))
        elif i[3]=='BIOSOLAR':
            schedBio.append(int(i[9]))
            metBio.append(int(i[10]))
        elif i[3]=='PERTALITE':
            schedPl.append(int(i[9]))
            metPl.append(int(i[10]))
    eoD = {'eodPREMIUM':sum(schedPre),
           'metPREMIUM':sum(metPre),
           'eodSOLAR':sum(schedSol),
           'metSOLAR':sum(metSol),
           'eodBIOSOLAR':sum(schedBio),
           'metBIOSOLAR':sum(metBio),
           'eodPERTAMAX':sum(schedPx),
           'metPERTAMAX':sum(metPx),
           'eodPERTALITE':sum(schedPl),
           'metPERTALITE':sum(metPl)}
    return eoD
def getExcelBLG(byMeter,byLo,bySched,dispet,patra,ptm):
    f = openpyxl.load_workbook(getSumTempSite('BLG'))
    lapSum = f.get_active_sheet()
    #thruput file
    g = openpyxl.load_workbook(getThruTempSite('BLG'))
    lapTru = g.get_active_sheet()
    #BYMETER PROCESSING
    meterUnionOp,meterUnionCl = getByMeterBLG(byMeter)
    #BYLO PROCESSING
    dLO = getDataLOBLG(byLo)
    #BYSCHEDULE PROCESSING
    eoD = getCarrierBLG(bySched)
    #SUMMARY INSERTION
    #BYMETER insertion to akhir meter
    #copy meterUnion opening totalizer awal
    for i, val in enumerate(meterUnionOp):
        lapSum.cell(row=i+7, column =4).value = val    
    #copy meterUnion closing to totalizer akhir
    for i, val in enumerate(meterUnionCl):
        lapSum.cell(row=i+7, column =5).value = val
    #BY LO insertion from DATA LO to kualitas lo fisik summary
    lapSum[7][7].value = dLO['PREMIUM'] #premium
    lapSum[14][7].value= dLO['SOLAR'] #solar
    lapSum[18][7].value= dLO['BIOSOLAR'] #biosolar spbu
    lapSum[21][7].value= dLO['BIOSOLAR PEMBANGKIT'] #biosolar pembangkit
    lapSum[26][7].value= dLO['PERTAMAX']  #pertamax
    lapSum[28][7].value= dLO['PERTALITE']  #pertalite
    lapSum[30][7].value= dLO['PERTAMAX TURBO'] #pertamax-turbo
    lapSum[36][7].value= dLO['AVTUR'] #avtur
    #BY Schedule insertion from Carrier to By schedule
    lapSum[7][8].value = eoD['eodPREMIUM'] #premium
    lapSum[14][8].value= eoD['eodSOLAR'] #solar
    lapSum[18][8].value= eoD['eodBIOSOLAR'] #biosolar spbu
    lapSum[21][8].value= eoD['eodBIOSOLAR-PL']#biosolar pembangkit
    lapSum[26][8].value= eoD['eodPERTAMAX']  #pertamax
    lapSum[28][8].value= eoD['eodPERTALITE']  #pertalite
    lapSum[30][8].value= eoD['eodPERTAMAX-TURBO'] #pertamax-turbo
    lapSum[36][8].value= eoD['eodAVTUR'] #avtur

    #THRUPUT BYMETER PROCESSING
    sumPre = []
    sumFame = []
    for i in range(7,13): 
        sumPre.append(lapSum.cell(row=i,column=5).value - lapSum.cell(row=i,column=4).value)
    for i in range(22,25):
        sumFame.append(lapSum.cell(row=i,column=5).value - lapSum.cell(row=i,column=4).value)
    #Minus operation because of none type
    sol1 = lapSum[13][4].value - lapSum[13][3].value
    sol2 = lapSum[17][4].value - lapSum[17][3].value
    sol3 = lapSum[19][4].value - lapSum[19][3].value
    sumSol = sol1 + sol2 + sol3
    #Minus operation because of none type
    px1 = lapSum[25][4].value - lapSum[25][3].value
    px2 = lapSum[27][4].value - lapSum[27][3].value
    px3 = lapSum[29][4].value - lapSum[29][3].value
    sumPx= px1+px2+px3
    #Minus operation because of single operation
    sumPxt = lapSum[30][4].value - lapSum[30][3].value
    #Minus operation because of single operation
    avt1 = lapSum[36][4].value - lapSum[36][3].value
    avt2 = lapSum[37][4].value - lapSum[37][3].value
    sumAvt = avt1 + avt2

    #THRUPUT INSERTION
    #By Meter
    lapTru[9][5].value = sum(sumPre) #premium
    lapTru[12][5].value= sumPx #pertamax
    lapTru[13][5].value= sumPxt#pertamax turbo
    lapTru[22][5].value= sumSol#solarTotal
    lapTru[23][5].value= sum(sumFame)#fame
    lapTru[6][6].value = eoD['metPERTALITE']#meterPertalite from carrier
    lapTru[16][6].value= eoD['metSOLAR']#meterSolar murni from carrier
    lapTru[19][6].value= eoD['metBIOSOLAR-PL']#meterBiosolarPL
    lapTru[14][5].value= sumAvt
    #By LO
    lapTru[6][3].value = dLO['PERTALITE']#pertalite
    lapTru[7][3].value = dLO['PREMIUM']#premium
    lapTru[10][3].value= dLO['PERTAMAX']#pertamax
    lapTru[13][3].value= dLO['PERTAMAX TURBO']#pertamax-turbo
    lapTru[16][3].value= dLO['SOLAR']#solar
    lapTru[18][3].value= dLO['BIOSOLAR']#biosolar
    lapTru[19][3].value= dLO['BIOSOLAR PEMBANGKIT']#biosolarPL
    lapTru[14][3].value= dLO['AVTUR']#avtur
    #By Scheduling
    lapTru[6][7].value =eoD['eodPERTALITE']#pertalite
    lapTru[7][7].value =eoD['eodPREMIUM']#premium
    lapTru[10][7].value=eoD['eodPERTAMAX']#pertamax
    lapTru[13][7].value=eoD['eodPERTAMAX-TURBO']#pertamax-turbo
    lapTru[16][7].value=eoD['eodSOLAR']#solar
    lapTru[18][7].value=eoD['eodBIOSOLAR']#biosolar
    lapTru[19][7].value=eoD['eodBIOSOLAR-PL']#biosolarPL
    lapTru[14][7].value=eoD['eodAVTUR']#avtur
    #FINISHING
    #1. CHANGE THE DATE
    todays = getFullDate()
    lapSum[4][0].value = todays
    lapTru[4][0].value = todays
    nowOn = 'Pukul : %s  WIB' % getHour()
    lapTru[26][0].value = nowOn
    #2. CHANGE THE NAME
    lapSum[50][1].value = dispet
    lapSum[50][6].value = patra
    lapSum[50][11].value = ptm
    lapTru[35][0].value = dispet
    lapTru[35][3].value = patra
    lapTru[35][9].value = ptm
    return f,g
def getExcelBYL(byMeter,byLo,bySched,fileSum,dispet,patra,ptm):
    f = openpyxl.load_workbook(fileSum)
    lapSum = f.get_active_sheet()
    #thruput file
    g = openpyxl.load_workbook(getThruTempSite('BYL'))
    lapTru = g.get_active_sheet()
    #BYMETER PROCESSING
    meterUnion = getByMeterBYL(byMeter)
    #BYLO PROCESSING
    dLO = getDataLOBYL(byLo)
    #BYSCHEDULE PROCESSING
    eoD = getCarrierBYL(bySched)
    # SUMMARY INSERTION
    # meter insertion to summary
    # copy paste totalizer akhir jadi awal
    for i in range(8, 31):
        lapSum.cell(row=i, column=4).value = lapSum.cell(row=i, column=5).value
    # copy meterUnion to totalizer akhir
    for i, val in enumerate(meterUnion):
        lapSum.cell(row=i + 8, column=5).value = val
    # BY LO insertion dari DATA LO to kualitas lo fisik summary
    lapSum[8][7].value = dLO['PREMIUM']#premium
    lapSum[18][7].value = dLO['SOLAR']#solar
    lapSum[20][7].value = dLO['BIOSOLAR']#biosolar
    lapSum[26][7].value = dLO['PERTAMAX']#pertamax
    lapSum[29][7].value = dLO['PERTALITE']#pertalite
    # BY Schedule insertion from Carrier
    lapSum[8][8].value = eoD['eodPREMIUM']#premium
    lapSum[18][8].value = eoD['eodSOLAR']#solar
    lapSum[20][8].value = eoD['eodBIOSOLAR']#biosolar
    lapSum[26][8].value = eoD['eodPERTAMAX']#pertamax
    lapSum[29][8].value = eoD['eodPERTALITE']#pertalite
    # THRUPUT BYMETER PROCESSING
    sumPre=[]
    sumPx=[]
    sumSol=[]
    sumFame=[]
    for i in range(8, 17):
        sumPre.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(25, 30):
        sumPx.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(17, 20):
        sumSol.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(21, 24):
        sumFame.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    # THRUPUT INSERTION
    # By meter
    lapTru[7][2].value = sum(sumPre)#premium
    lapTru[11][2].value = sum(sumPx)#pertamax
    lapTru[13][2].value = eoD['metPERTALITE']#pertalite dari metPl
    lapTru[17][2].value = sum(sumSol)#solar
    lapTru[19][2].value = eoD['metBIOSOLAR']#biosolar daru metBio
    lapTru[21][2].value = sum(sumFame)#fame
    # By Schedule
    lapTru[5][1].value = eoD['eodPREMIUM']#premium
    lapTru[9][1].value = eoD['eodPERTAMAX']#pertamax
    lapTru[13][1].value = eoD['eodPERTALITE']#pertalite
    lapTru[15][1].value = eoD['eodSOLAR']#solar
    lapTru[19][1].value = eoD['eodBIOSOLAR']#biosolar
    # By LO
    lapTru[5][3].value = dLO['PREMIUM']#premium
    lapTru[9][3].value = dLO['PERTAMAX']#pertamax
    lapTru[13][3].value = dLO['PERTALITE']#pertalite
    lapTru[15][3].value = dLO['SOLAR']#solar
    lapTru[19][3].value = dLO['BIOSOLAR']#biosolar
    # FINISHING
    # 1. CHANGE THE DATE
    todays = getDateDash()
    thisDay = 'DATE : %s' % getDateDash()
    lapSum[3][0].value = todays
    lapTru[1][0].value = thisDay
    # 2. CHANGE THE NAME
    lapSum[39][0].value = dispet
    lapSum[39][7].value = patra
    lapSum[39][11].value = ptm
    lapTru[31][0].value = dispet
    lapTru[31][2].value = patra
    lapTru[31][4].value = ptm
    return f,g
def getExcelKTP(byMeter,byLo,bySched,fileSum,dispet,patra,ptm,*args):
    f = openpyxl.load_workbook(fileSum)
    lapSum = f.get_active_sheet()
    #thruput file
    g = openpyxl.load_workbook(getThruTempSite('KTP'))
    lapTru = g.get_active_sheet()
    #BYMETER PROCESSING
    meterUnion = getByMeterKTP(byMeter)
    #BYLO PROCESSING
    dLO = getDataLOKTP(byLo)
    #BYSCHEDULE PROCESSING
    eoD = getCarrierKTP(bySched)
    #SUMMARY INSERTION
    #meter insertion to akhir meter
    #copy paste totalizer akhir jadi awal
    for i in range(11,42):
        lapSum.cell(row=i, column=4).value =  lapSum.cell(row=i, column=5).value
    #copy meterUnion to totalizer akhir
    for i, val in enumerate(meterUnion):
        lapSum.cell(row=i+11, column =5).value = val
    # BY LO insertion from DATA LO to kualitas lo fisik summary
    lapSum[11][9].value = '=i11'#premium
    lapSum[18][9].value = '=i18'#PremiumRTW dan PremiumTKG
    lapSum[20][9].value = '=i20'#solarMurni
    lapSum[22][9].value = dLO['BIOSOLAR']#biosolarSpbu
    lapSum[24][9].value = dLO['BIOSOLAR-PL']#biosolarPembangkit
    lapSum[26][9].value = '=i26'#SolarRTW dan SolarTKG
    lapSum[35][9].value = dLO['PERTAMAX']#pertamax
    lapSum[37][9].value = dLO['PERTALITE']#pertalite
    lapSum[39][9].value = dLO['PERTAMINA-DEX']#pertaminaDex
    lapSum[41][9].value = dLO['DEXLITE']#dexlite
    # BY Schedule insertion from Carrier to By schedule
    lapSum[11][8].value = eoD['eodPREMIUM']  # premium
    lapSum[18][8].value = eoD['eodPREMIUM-RTW-TKG']  # PremiumRTW dan PremiumTKG
    lapSum[20][8].value = eoD['eodSOLAR']  # solarMurni
    lapSum[22][8].value = eoD['eodBIOSOLAR']# biosolarSpbu
    lapSum[24][8].value = eoD['eodBIOSOLAR-PL']# biosolarPembangkit
    lapSum[26][8].value = eoD['eodSOLAR-RTW-TKG']  # SolarRTW dan SolarTKG
    lapSum[35][8].value = eoD['eodPERTAMAX']# pertamax
    lapSum[37][8].value = eoD['eodPERTALITE']# pertalite
    lapSum[39][8].value = eoD['eodDEX']# pertaminaDex
    lapSum[41][8].value = eoD['eodDEXLITE']# dexlite
    # THRUPUT BYMETER PROCESSING
    sumPre = []
    sumSol = []
    sumFame= []
    for i in range(11, 19):
        sumPre.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(19, 23):
        sumSol.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(24, 27):  # check if error
        sumSol.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(27, 32):
        sumFame.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    fame1 = lapSum[33][4].value - lapSum[33][3].value
    sumFame.append(fame1)
    # Minus operation because of none type
    px1 = lapSum[34][4].value - lapSum[34][3].value
    px2 = lapSum[36][4].value - lapSum[36][3].value
    sumPx = px1 + px2
    # Minus operation because of single operation
    sumDex = lapSum[38][4].value - lapSum[38][3].value
    # THRUPUT INSERTION
    # By Meter
    lapTru[12][2].value = eoD['metPREMIUM-RTW-TKG']  # premiumRTW dan premiumTKG
    lapTru[13][2].value = sum(sumPre)  # premium
    lapTru[15][2].value = eoD['metPERTAMAX']# meterpertamax
    lapTru[17][2].value = sumPx# Pertamax
    lapTru[19][2].value = eoD['metPERTALITE']# pertalite
    lapTru[24][2].value = eoD['metSOLAR-RTW-TKG']  # solarRTW dan solarTKG
    lapTru[26][2].value = sum(sumSol)  # Solar
    lapTru[28][2].value = eoD['metBIOSOLAR']# meterBiosolar
    lapTru[29][2].value = eoD['metBIOSOLAR-PL']# meterBiosolarPL
    lapTru[33][2].value = sum(sumFame)  # fame
    lapTru[35][2].value = eoD['metDEX']# meterDex
    lapTru[37][2].value = sumDex  # Dex
    lapTru[39][2].value = eoD['metDEXLITE']# meterDexlite
    # By LO
    lapTru[10][3].value = '=b10'  # premium
    lapTru[12][3].value = '=b12'  # premiumRTW dan premiumTKG
    lapTru[15][3].value = dLO['PERTAMAX']  # pertamax
    lapTru[19][3].value = dLO['PERTALITE']  # pertalite
    lapTru[21][3].value = '=b21'  # solar
    lapTru[23][3].value = '=b23'  # solar pembangkit ############
    lapTru[24][3].value = '=b24'  # solarRTW dan solarTKG
    lapTru[28][3].value = dLO['BIOSOLAR']# biosolar
    lapTru[29][3].value = dLO['BIOSOLAR-PL']  # biosolarPL
    lapTru[35][3].value = dLO['PERTAMINA-DEX']  # Dex
    lapTru[39][3].value = dLO['DEXLITE']  # dexlite
    # By Scheduling
    lapTru[10][1].value = eoD['eodPREMIUM']# premium
    lapTru[12][1].value = eoD['eodPREMIUM-RTW-TKG']# premiumRTW dan premiumTKG
    lapTru[15][1].value = eoD['eodPERTAMAX']# pertamax
    lapTru[19][1].value = eoD['eodPERTALITE']# pertalite
    lapTru[21][1].value = eoD['eodSOLAR']# solar
    lapTru[23][1].value = eoD['eodSOLAR-PL']# solar pembangkit
    lapTru[24][1].value = eoD['eodSOLAR-RTW-TKG']# solarRTW dan solarTKG
    lapTru[28][1].value = eoD['eodBIOSOLAR']# biosolar
    lapTru[29][1].value = eoD['eodBIOSOLAR-PL']# biosolarPL
    lapTru[35][1].value = eoD['eodDEX']# Dex
    lapTru[39][1].value = eoD['eodDEXLITE']# dexlite
    # FINISHING
    # 1. CHANGE THE DATE
    thisDay = getDateDash()
    todays = 'DATE : %s' % thisDay
    nowOn = getDate()
    lapSum[7][0].value = thisDay
    lapTru[5][0].value = todays
    # 2. CHANGE THE NAME
    lapSum[51][1].value = dispet
    lapSum[51][6].value = patra
    lapSum[51][9].value = ptm
    lapSum[51][11].value = args[0]
    lapTru[48][0].value = dispet
    lapTru[48][1].value = patra
    lapTru[48][2].value = ptm
    lapTru[48][4].value = args[0]
    return f,g
def getExcelMDN(byMeter,byLo,bySched,fileSum,dispet,patra,ptm):
    f = openpyxl.load_workbook(fileSum)
    lapSum = f.get_active_sheet()
    #thruput file
    g = openpyxl.load_workbook(getThruTempSite('MDN'))
    lapTru = g.get_active_sheet()
    #BYMETER PROCESSING
    meterUnion = getByMeterMDN(byMeter)
    #BYLO PROCESSING
    dLO = getDataLOMDN(byLo)
    #BYSCHEDULE PROCESSING
    eoD = getCarrierMDN(bySched)
    # meter insertion to summary
    # copy paste totalizer akhir jadi awal
    for i in range(7, 43):
        lapSum.cell(row=i, column=4).value = lapSum.cell(row=i, column=5).value
    # copy meterUnion to totalizer akhir
    for i, val in enumerate(meterUnion):
        lapSum.cell(row=i + 7, column=5).value = val
        # BY LO insertion from ByLoProcessing
    lapSum[8][7].value = dLO['PREMIUM']  # premium
    lapSum[20][7].value = dLO['SOLAR']  # solar
    lapSum[22][7].value = dLO['SOLAR RTW'] # solarRtw
    lapSum[26][7].value = dLO['DEXLITE']  # dexlite
    lapSum[28][7].value = dLO['BIOSOLAR']  # biosolar
    lapSum[35][7].value = dLO['PERTAMAX']  # pertamax
    lapSum[38][7].value = dLO['PERTALITE']  # pertalite
    lapSum[41][7].value = dLO['PERTALITE RTW']  # pertaliteRtw

    # BY Schedule insertion from Carrier to By schedule
    lapSum[8][8].value = eoD['eodPREMIUM']# premium
    lapSum[20][8].value = eoD['eodSOLAR']# solar
    lapSum[22][8].value = eoD['eodSOLAR-RTW']# solarRtw
    lapSum[26][8].value = eoD['eodDEXLITE']# dexlite
    lapSum[28][8].value = eoD['eodBIOSOLAR']# biosolar
    lapSum[35][8].value = eoD['eodPERTAMAX']# pertamax
    lapSum[38][8].value = eoD['eodPERTALITE']# pertalite
    lapSum[41][8].value = eoD['eodPERTALITE-RTW']# pertaliteRtw

    # THRUPUT BYMETER PROCESSING
    sumPre  = []
    sumPrePl= []
    sumFame = []
    sumPx = []
    for i in range(7, 11):
        sumPre.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(11, 19):
        sumPrePl.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    # Minus operation because of none type
    sol1 = lapSum[19][4].value - lapSum[19][3].value
    sol2 = lapSum[21][4].value - lapSum[21][3].value
    sol3 = lapSum[23][4].value - lapSum[23][3].value
    sol4 = lapSum[24][4].value - lapSum[24][3].value
    sol5 = lapSum[26][4].value - lapSum[26][3].value
    sol6 = lapSum[27][4].value - lapSum[27][3].value
    sol7 = lapSum[28][4].value - lapSum[28][3].value
    sumSol = sol1 + sol2 + sol3 + sol4 + sol5 + sol6 + sol7
    for i in range(29, 34):
        sumFame.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(34, 42):
        sumPx.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)

    # THRUPUT INSERTION
    # By Meter
    lapTru[7][2].value = sum(sumPre)  # premium
    lapTru[10][2].value = sum(sumPx)  # pertamax
    lapTru[22][2].value = sumSol  # solarTotal
    lapTru[21][2].value = sum(sumFame)  # fame
    lapTru[11][2].value = eoD['metPERTALITE-RTW']# meterPertalite from carrier
    lapTru[13][2].value = eoD['metSOLAR']# meterSolar murni from carrier
    lapTru[12][2].value = eoD['metPERTALITE']- eoD['metPERTALITE-RTW']
    lapTru[14][2].value = eoD['metSOLAR-RTW']
    lapTru[15][2].value = eoD['metDEXLITE']
    # By LO
    lapTru[5][3].value = dLO['PREMIUM']  # premium
    lapTru[8][3].value = dLO['PERTAMAX']  # pertamax
    lapTru[11][3].value = dLO['PERTALITE RTW']  # pertaliteRTW
    lapTru[12][3].value = dLO['PERTALITE']  # pertalite
    lapTru[13][3].value = dLO['SOLAR']  # solar
    lapTru[14][3].value = dLO['SOLAR RTW']  # solRTW
    lapTru[15][3].value = dLO['DEXLITE']  # Dexlite
    lapTru[17][3].value = dLO['BIOSOLAR']  # biosolar
    # By Scheduling
    lapTru[5][1].value = eoD['eodPREMIUM']  # premium
    lapTru[8][1].value = eoD['eodPERTAMAX']  # pertamax
    lapTru[11][1].value = eoD['eodPERTALITE-RTW']  # pertaliteRTW
    lapTru[12][1].value = eoD['eodPERTALITE']  # pertalite
    lapTru[13][1].value = eoD['eodSOLAR']  # solar
    lapTru[14][1].value = eoD['eodSOLAR-RTW']  # solarRTW
    lapTru[15][1].value = eoD['eodDEXLITE']  # dexlite
    lapTru[17][1].value = eoD['eodBIOSOLAR'] # biosolar

    # FINISHING
    # 1. CHANGE THE DATE
    todays = getFullDate
    thisDay = 'DATE : %s' %getDateDash
    lapSum[4][0].value = todays
    lapTru[1][0].value = thisDay
    # 2. CHANGE THE NAME
    lapSum[51][0].value = dispet
    lapSum[51][6].value = patra
    lapSum[51][11].value = ptm
    lapTru[33][0].value = dispet
    lapTru[33][2].value = patra
    lapTru[33][4].value = ptm
    return f,g
def getExcelPLP(byMeter,byLo,bySched,fileSum,dispet,patra,ptm,*args):
    f = openpyxl.load_workbook(fileSum)
    lapSum = f.get_active_sheet()
    #thruput file
    g = openpyxl.load_workbook(getThruTempSite('PLP'))
    lapTru = g.get_active_sheet()
    #BYMETER PROCESSING
    akhirMeter = getByMeterPLP(byMeter)
    #BYLO PROCESSING
    dLO = getDataLOPLP(byLo)
    #BYSCHEDULE PROCESSING
    eoD = getCarrierPLP(bySched)
    # awal jadi akhir
    for i in range(6, 67):
        lapSum.cell(row=i, column=4).value = lapSum.cell(row=i, column=5).value
    # csv to xls
    for i, val in enumerate(akhirMeter):
        lapSum.cell(row=i + 6, column=5).value = val
    # by LO
    lapSum[7][9].value = dLO['PREMIUM']#PREMIUM
    lapSum[18][9].value = dLO['PERTALITE']#PERTALITE
    lapSum[28][9].value = dLO['PERTAMAX']#PERTAMAX
    lapSum[41][9].value = dLO['SOLAR']#SOLAR
    lapSum[44][9].value = dLO['BIOSOLAR']#BIOSOLAR
    lapSum[60][9].value = dLO['PERTAMAX TURBO']#PERTAMAX-TURBO
    lapSum[62][9].value = dLO['DEXLITE']#DEXLITE
    lapSum[64][9].value = dLO['PERTAMINA DEX']#PERTAMINA-DEX
    # eod to lapsum by sched
    lapSum[7][8].value = eoD['eodPREMIUM']#PREMIUM
    lapSum[18][8].value = eoD['eodPERTALITE']#PERTALITE
    lapSum[28][8].value = eoD['eodPERTAMAX']#PERTAMAX
    lapSum[41][8].value = eoD['eodSOLAR']#SOLAR
    lapSum[44][8].value = eoD['eodBIOSOLAR']#BIOSOLAR
    lapSum[60][8].value = eoD['eodPERTAMAX-TURBO']#PERTAMAX-TURBO
    lapSum[62][8].value = eoD['eodDEXLITE']#DEXLITE
    lapSum[64][8].value = eoD['eodDEX']#PERTAMINA-DEX
    # Laporan Thruput
    # meter Summary
    sumPre = []
    sumPx = []
    sumSol = []
    sumFame = []
    sumPxT = []
    sumDex = []
    for i in range(6, 27):  # 28
        sumPre.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(27, 40):  # 28,40
        sumPx.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(40, 50):
        sumSol.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(50, 59):
        sumFame.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(59, 61):
        sumPxT.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    dex1 = lapSum[61][4].value - lapSum[61][3].value
    dex2 = lapSum[63][4].value - lapSum[63][3].value
    sumDex = dex1 + dex2

    # byScheduling
    lapTru[5][1].value = eoD['eodPREMIUM']#PREMIUM
    lapTru[6][1].value = eoD['eodSOLAR']#SOLAR
    lapTru[7][1].value = eoD['eodBIOSOLAR']#BIOSOLAR
    lapTru[10][1].value = eoD['eodDEXLITE']#DEXLITE
    lapTru[16][1].value = eoD['eodPERTAMAX']#PERTAMAX
    lapTru[17][1].value = eoD['eodPERTALITE']#PERTALITE
    lapTru[18][1].value = eoD['eodPERTAMAX-TURBO']#PERTAMAX-TURBO
    lapTru[19][1].value = eoD['eodDEX']#PERTAMINA-DEX
    # byMeter
    lapTru[6][2].value = eoD['metSOLAR']#SOLAR
    lapTru[17][2].value = eoD['metPERTALITE']#PERTALITE
    lapTru[10][2].value = eoD['metDEXLITE']#DEXLITE
    lapTru[5][2].value = '=%i-((45/100)*C17)' % sum(sumPre)#PREMIUM
    lapTru[16][2].value = '=%i-((55/100)*C17)' % sum(sumPx)#PERTAMAX
    lapTru[12][2].value = sum(sumFame)#FAME
    lapTru[15][2].value = sum(sumSol)#SOLAR TOTAL
    lapTru[18][2].value = sum(sumPxT)#PERTAMAX-TURBO
    lapTru[20][2].value = '=20/100*C10'
    lapTru[19][2].value = '=%i-C20' % sumDex #PERTAMINA-DEX
    # by LO
    lapTru[5][3].value = dLO['PREMIUM']#PREMIUM
    lapTru[6][3].value = dLO['SOLAR']#SOLAR
    lapTru[7][3].value = dLO['BIOSOLAR']#BIOSOLAR
    lapTru[10][3].value = dLO['DEXLITE']#DEXLITE
    lapTru[16][3].value = dLO['PERTAMAX']#PERTAMAX
    lapTru[17][3].value = dLO['PERTALITE']#PERTALITE
    lapTru[18][3].value = dLO['PERTAMAX TURBO']#PERTAMAX-TURBO
    lapTru[19][3].value = dLO['PERTAMINA DEX']#PERTAMINA-DEX
    lapTru['D11'] = '=0.8*D7'
    lapTru['D15'] = '=D11+D6+D14'

    todays = getDashDate()
    lapSum[2][0].value = todays
    lapTru['a1'].value = 'Date : %s' % todays
    lapSum[75][1].value = dispet
    lapSum[75][5].value = patra
    lapSum[75][8].value = ptm
    lapSum[75][10].value = args[0]
    lapTru['a30'].value = dispet
    lapTru['b30'].value = patra
    lapTru['c30'].value = ptm
    lapTru['e30'].value = args[0]
    return f,g
def getExcelSBY(byMeter,byLo,bySched,fileSum,dispet,patra,ptm):
    f = openpyxl.load_workbook(fileSum)
    lapSum = f.get_active_sheet()
    #thruput file
    g = openpyxl.load_workbook(getThruTempSite('SBY'))
    lapTru = g.get_active_sheet()
    #BYMETER PROCESSING
    meterUnion = getByMeterSBY(byMeter)
    #BYLO PROCESSING
    dLO = getDataLOSBY(byLo)
    #BYSCHEDULE PROCESSING
    eoD = getCarrierSBY(bySched)
    # SUMMARY INSERTION
    # meter insertion to akhir meter
    # copy paste totalizer akhir jadi awal
    for i in range(6, 67):
        lapSum.cell(row=i, column=4).value = lapSum.cell(row=i, column=5).value
    # copy meterUnion to totalizer akhir
    for i, val in enumerate(meterUnion):
        lapSum.cell(row=i + 6, column=5).value = val
    # BY LO insertion from DATA LO to kualitas lo fisik summary
    lapSum[6][8].value = dLO['PREMIUM']#premium
    lapSum[19][8].value = dLO['PERTALITE']#pertalite
    lapSum[30][8].value = dLO['SOLAR']#solar murni
    lapSum[34][8].value = dLO['BIOSOLAR']#biosolar
    lapSum[37][8].value = dLO['DEXLITE']# dexlite
    lapSum[41][8].value = dLO['BIOSOLAR PEMBANGKIT']# biosolarPL
    lapSum[54][8].value = dLO['PERTAMAX']# pertamax
    lapSum[63][8].value = dLO['PERTAMAX TURBO']# pertamax-turbo
    lapSum[64][8].value = dLO['PERTAMINA DEX']# pertaminaDex

    # BY Schedule insertion from Carrier to By schedule
    lapSum[6][7].value = eoD['eodPREMIUM']# premium
    lapSum[19][7].value = eoD['eodPERTALITE']  # pertalite
    lapSum[30][7].value = eoD['eodSOLAR']  # solar murni
    lapSum[34][7].value = eoD['eodBIOSOLAR']  # biosolar
    lapSum[37][7].value = eoD['eodDEXLITE']  # dexlite
    lapSum[41][7].value = eoD['eodBIOSOLAR-PL']  # biosolarPl
    lapSum[54][7].value = eoD['eodPERTAMAX']  # pertamax
    lapSum[63][7].value = eoD['eodPERTAMAX-TURBO']  # pertamax-turbo
    lapSum[64][7].value = eoD['eodDEX']  # pertaminaDex

    # THRUPUT BYMETER PROCESSING
    sumPre = []
    sumPrePl=[]
    sumSol = []
    sumFame = []
    sumPx = []
    sumPxt = []
    sumDex= []
    for i in range(6, 11):
        sumPre.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(12, 15):
        sumPre.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(15, 25):
        sumPrePl.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    # Minus operation because of none type
    sol1 = lapSum[29][4].value - lapSum[29][3].value
    sol2 = lapSum[33][4].value - lapSum[33][3].value
    sol3 = lapSum[35][4].value - lapSum[35][3].value
    sol4 = lapSum[37][4].value - lapSum[37][3].value
    sol5 = lapSum[39][4].value - lapSum[39][3].value
    sol6 = lapSum[41][4].value - lapSum[41][3].value
    sol7 = lapSum[43][4].value - lapSum[43][3].value
    sumSol = sol1 + sol2 + sol3 + sol4 + sol5 + sol6 + sol7
    for i in range(45, 52):
        sumFame.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(52, 63):
        sumPx.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    # Minus operation because of single operation
    sumPxt = lapSum[63][4].value - lapSum[63][3].value
    # Minus operation because of none type
    dex1 = lapSum[64][4].value - lapSum[64][3].value
    dex2 = lapSum[66][4].value - lapSum[66][3].value
    sumDex = dex1 + dex2

    # THRUPUT INSERTION
    # By Meter
    lapTru[5][2].value = sum(sumPre) + sum(sumPrePl)  # premium
    lapTru[11][2].value = eoD['metSOLAR']  # solar
    lapTru[13][2].value = eoD['metBIOSOLAR']  # biosolar
    lapTru[16][2].value = eoD['metBIOSOLAR-PL']  # biosolar PBL
    lapTru[18][2].value = sumSol  # solar total
    lapTru[20][2].value = sum(sumFame)  # fame
    lapTru[23][2].value = sum(sumPx)  # pertamax
    lapTru[26][2].value = eoD['metPERTALITE']  # pertalite
    lapTru[27][2].value = eoD['metDEXLITE']  # dexlite
    lapTru[28][2].value = sumPxt  # pertamax-turbo
    lapTru[29][2].value = sumDex  # pertamina-dex
    # By LO
    lapTru[5][3].value = dLO['PREMIUM']  # premium
    lapTru[11][3].value = dLO['SOLAR']  # solar
    lapTru[13][3].value = dLO['BIOSOLAR']  # biosolar
    lapTru[16][3].value = dLO['BIOSOLAR PEMBANGKIT']  # biosolar PL
    lapTru[23][3].value = dLO['PERTAMAX']  # pertamax
    lapTru[26][3].value = dLO['PERTALITE']  # pertalite
    lapTru[27][3].value = dLO['DEXLITE']  # dexlite
    lapTru[28][3].value = dLO['PERTAMAX TURBO']  # pertamax-Turbo
    lapTru[29][3].value = dLO['PERTAMINA DEX']  # pertaminaDex
    # By Scheduling
    lapTru[5][1].value = eoD['eodPREMIUM']  # premium
    lapTru[11][1].value = eoD['eodSOLAR'] # solar
    lapTru[13][1].value = eoD['eodBIOSOLAR']  # biosolar
    lapTru[16][1].value = eoD['eodBIOSOLAR-PL']  # biosolarPL
    lapTru[23][1].value = eoD['eodPERTAMAX']  # pertamax
    lapTru[26][1].value = eoD['eodPERTALITE']  # pertalite
    lapTru[27][1].value = eoD['eodDEXLITE']  # dexlite
    lapTru[28][1].value = eoD['eodPERTAMAX-TURBO']  # pertamax-turbo
    lapTru[29][1].value = eoD['eodDEX']  # pertamina-dex
    # FINISHING
    # 1. CHANGE THE DATE
    todays = getDateDash()
    thisDay = 'DATE : %s'% getFullDate()
    lapSum[2][0].value = todays
    lapTru[1][0].value = thisDay
    # 2. CHANGE THE NAME
    lapSum[74][1].value = dispet
    lapSum[74][6].value = patra
    lapSum[74][8].value = ptm
    lapTru[39][0].value = dispet
    lapTru[39][2].value = patra
    lapTru[39][4].value = ptm
    return f,g
def getExcelTGR(byMeter,byLo,bySched,fileSum,dispet,patra,ptm):
    f = openpyxl.load_workbook(fileSum)
    lapSum = f.get_active_sheet()
    #thruput file
    g = openpyxl.load_workbook(getThruTempSite('TGR'))
    lapTru = g.get_active_sheet()
    #BYMETER PROCESSING
    meterUnion = getByMeterTGR(byMeter)
    #BYLO PROCESSING
    dLO = getDataLOTGR(byLo)
    #BYSCHEDULE PROCESSING
    eoD = getCarrierTGR(bySched)
    #meter insertion to summary
    #copy paste totalizer akhir jadi awal
    for i in range(7,36):
        lapSum.cell(row=i, column=4).value =  lapSum.cell(row=i, column=5).value
    #copy meterUnion to totalizer akhir
    for i, val in enumerate(meterUnion):
        lapSum.cell(row=i+7, column =5).value = val
    #BY LO insertion dari DATA LO to kualitas lo fisik summary
    lapSum[7][7].value =dLO['PREMIUM'] #premium
    lapSum[14][7].value=dLO['SOLAR'] #solar
    lapSum[18][7].value=dLO['BIOSOLAR'] #biosolar
    lapSum[20][7].value=dLO['BIOSOLAR PEMBANGKIT'] #biosolar pembangkit
    lapSum[26][7].value=dLO['PERTAMAX']#pertamax
    lapSum[28][7].value=dLO['PERTALITE']#pertalite
    lapSum[30][7].value=dLO['MFO']#MFO
    #BY Schedule insertion from Carrier
    lapSum[7][8].value = eoD['eodPREMIUM'] #premium
    lapSum[14][8].value= eoD['eodSOLAR']#solar
    lapSum[18][8].value= eoD['eodBIOSOLAR'] #biosolar
    lapSum[20][8].value= eoD['eodBIOSOLAR-PL']#biosolar pembangkit
    lapSum[26][8].value= eoD['eodPERTAMAX']#pertamax
    lapSum[28][8].value= eoD['eodPERTALITE']#pertalite
    lapSum[30][8].value= eoD['eodMFO']#MFO
    #THRUPUT BYMETER PROCESSING
    sumPre = []
    sumFame= []
    for i in range(7,13):
        sumPre.append(lapSum.cell(row=i,column=5).value - lapSum.cell(row=i,column=4).value)
    for i in range(21,25):
        sumFame.append(lapSum.cell(row=i,column=5).value - lapSum.cell(row=i,column=4).value)
    #Minus operation because of none type
    sol1 = lapSum[13][4].value - lapSum[13][3].value
    sol2 = lapSum[15][4].value - lapSum[15][3].value
    sol3 = lapSum[17][4].value - lapSum[17][3].value
    sol4 = lapSum[19][4].value - lapSum[19][3].value
    sol5 = lapSum[20][4].value - lapSum[20][3].value
    sumSol = sol1 + sol2 + sol3 + sol4 + sol5
    #Minus operation because of none type
    px1 = lapSum[25][4].value - lapSum[25][3].value
    px2 = lapSum[26][4].value - lapSum[26][3].value
    px3 = lapSum[27][4].value - lapSum[27][3].value
    sumPx = px1 + px2 + px3
    #MFO
    sumMFO = lapSum[29][4].value - lapSum[29][3].value
    #THRUPUT INSERTION
    #By Meter
    lapTru[6][6].value = eoD['metPERTALITE']#pertalite
    lapTru[9][5].value= sum(sumPre) #premium
    lapTru[12][5].value= sumPx #pertamax
    lapTru[17][5].value= sumMFO #MFO
    lapTru[24][5].value = sumSol #solar meter
    lapTru[25][5].value= sum(sumFame) #fame
    lapTru[18][6].value= eoD['metSOLAR']
    lapTru[21][6].value= eoD['metBIOSOLAR-PL']
    #By LO
    lapTru[6][3].value = dLO['PERTALITE']#pertalite
    lapTru[7][3].value = dLO['PREMIUM']#premium
    lapTru[10][3].value= dLO['PERTAMAX']#pertamax
    lapTru[13][3].value= dLO['MFO']#mfo
    lapTru[18][3].value= dLO['SOLAR']#solar
    lapTru[20][3].value= dLO['BIOSOLAR']#biosolar
    lapTru[21][3].value= dLO['BIOSOLAR PEMBANGKIT']#biosolarPembangkit
    #By Scheduling
    lapTru[6][7].value =eoD['eodPERTALITE']#pertalite
    lapTru[7][7].value =eoD['eodPREMIUM']#premium
    lapTru[10][7].value=eoD['eodPERTAMAX']#pertamax
    lapTru[13][7].value=eoD['eodMFO']#mfo
    lapTru[18][7].value=eoD['eodSOLAR']#solar
    lapTru[20][7].value=eoD['eodBIOSOLAR']#Biosolar
    lapTru[21][7].value=eoD['eodBIOSOLAR-PL']#biosolarPembangkit
    #FINISHING
    #1. CHANGE THE DATE
    todays = getFullDate()
    lapSum[4][0].value = todays
    lapTru[4][0].value = todays
    nowOn = 'Pukul : %s' %getHour
    lapTru[28][0].value = nowOn
    #2. CHANGE THE NAME
    lapSum[44][1].value = dispet
    lapSum[44][6].value = patra
    lapSum[44][11].value = ptm
    lapTru[37][0].value = dispet
    lapTru[37][3].value = patra
    lapTru[37][9].value = ptm
    return f,g
def getExcelUJB(byMeter,byLo,bySched,fileSum,patra,ptm):
    f = openpyxl.load_workbook(fileSum)
    lapSum = f.get_active_sheet()
    #thruput file
    g = openpyxl.load_workbook(getThruTempSite('UJB'))
    lapTru = g.get_active_sheet()
    #BYMETER PROCESSING
    meterUnion = getByMeterUJB(byMeter)
    #BYLO PROCESSING
    dLO = getDataLOUJB(byLo)
    #BYSCHEDULE PROCESSING
    eoD = getCarrierUJB(bySched)
    # SUMMARY INSERTION
    # meter insertion to akhir meter
    # copy paste totalizer akhir jadi awal
    for i in range(6, 34):
        lapSum.cell(row=i, column=4).value = lapSum.cell(row=i, column=5).value
    # copy meterUnion to totalizer akhir
    for i, val in enumerate(meterUnion):
        lapSum.cell(row=i + 6, column=5).value = val
    # BY LO insertion from DATA LO to kualitas lo fisik summary
    lapSum[6][9].value = dLO['PREMIUM']  # premium
    lapSum[17][9].value = dLO['SOLAR']  # solar
    lapSum[21][9].value = dLO['BIOSOLAR']# biosolar
    lapSum[30][9].value = dLO['PERTAMAX']# pertamax
    lapSum[32][9].value = dLO['PERTALITE']  # pertalite
    # BY Schedule insertion from Carrier to By schedule
    lapSum[6][8].value = eoD['eodPREMIUM']  # premium
    lapSum[17][8].value = eoD['eodSOLAR']  # solar
    lapSum[21][8].value = eoD['eodBIOSOLAR']  # biosolar
    lapSum[30][8].value = eoD['eodPERTAMAX']  # pertamax
    lapSum[32][8].value = eoD['eodPERTALITE']  # pertalite

    # THRUPUT BYMETER PROCESSING
    sumPre = []
    sumPx = []
    for i in range(6, 16):
        sumPre.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    for i in range(29, 32):
        sumPx.append(lapSum.cell(row=i, column=5).value - lapSum.cell(row=i, column=4).value)
    # Minus operation because of none type
    sol1 = lapSum[16][4].value - lapSum[16][3].value
    sol2 = lapSum[18][4].value - lapSum[18][3].value
    sol3 = lapSum[20][4].value - lapSum[20][3].value
    sol4 = lapSum[21][4].value - lapSum[21][3].value
    sumSol = sol1 + sol2 + sol3 + sol4
    # Minus operation because of none type
    fame1 = lapSum[22][4].value - lapSum[22][3].value
    fame2 = lapSum[24][4].value - lapSum[24][3].value
    fame3 = lapSum[25][4].value - lapSum[25][3].value
    fame4 = lapSum[27][4].value - lapSum[27][3].value
    sumFame = fame1 + fame2 + fame3 + fame4

    # THRUPUT INSERTION
    # By Meter
    lapTru[7][2].value = sum(sumPre)  # premium
    lapTru[10][2].value = sum(sumPx)  # pertamax
    lapTru[11][2].value = eoD['metPERTALITE'] # pertalite
    lapTru[18][2].value = sumFame  # fame
    lapTru[19][2].value = sumSol  # solar
    lapTru[12][2].value = eoD['metSOLAR']# meterSolar murni from carrier
    # By LO
    lapTru[5][3].value = dLO['PREMIUM']  # premium
    lapTru[8][3].value = dLO['PERTAMAX']  # pertamax
    lapTru[11][3].value = dLO['PERTALITE']# pertalite
    lapTru[12][3].value = dLO['SOLAR']# solar
    lapTru[14][3].value = dLO['BIOSOLAR']# biosolar
    # By Scheduling
    lapTru[5][1].value = eoD['eodPREMIUM']  # premium
    lapTru[8][1].value = eoD['eodPERTAMAX']# pertamax
    lapTru[11][1].value = eoD['eodPERTALITE']# pertalite
    lapTru[12][1].value = eoD['eodSOLAR']# solar
    lapTru[14][1].value = eoD['eodBIOSOLAR']  # biosolar

    # FINISHING
    # 1. CHANGE THE DATE
    todays = getDateSlash()
    thisDay = 'DATE: %s' % getDashDate()
    lapSum[2][0].value = todays
    lapTru[1][0].value = thisDay
    # 2. CHANGE THE NAME
    lapSum[45][5].value = ptm
    lapSum[45][8].value = patra
    lapTru[29][1].value = ptm
    lapTru[29][3].value = patra
    return f,g
def mainProg(site,byMeter,byLo,bySched,dispet,patra,ptm):
    if site == 'BLG':
        summary,thruput = getExcelBLG(byMeter,byLo,bySched,dispet,patra,ptm)
        return summary,thruput
def mainProgDet(site,byMeter,byLo,bySched,fileSum,dispet,patra,ptm,*args):
    if site == 'BYL':
        summary,thruput = getExcelBYL(byMeter,byLo,bySched,fileSum,dispet,patra,ptm)
        return summary,thruput
    elif site == 'KTP':
        summary,thruput = getExcelKTP(byMeter,byLo,bySched,fileSum,dispet,patra,ptm,*args)
        return summary, thruput
    elif site == 'MDN':
        summary,thruput = getExcelMDN(byMeter,byLo,bySched,fileSum,dispet,patra,ptm)
        return summary, thruput
    elif site == 'PLP':
        summary,thruput = getExcelPLP(byMeter,byLo,bySched,fileSum,dispet,patra,ptm,*args)
        return summary, thruput
    elif site == 'SBY':
        summary,thruput = getExcelSBY(byMeter,byLo,bySched,fileSum,dispet,patra,ptm)
        return summary, thruput
    elif site == 'TGR':
        summary,thruput = getExcelTGR(byMeter,byLo,bySched,fileSum,dispet,patra,ptm)
        return summary, thruput
def mainProgEgo(site,byMeter,byLo,bySched,fileSum,patra,ptm):
    if site == 'UJB':
        summary,thruput = getExcelUJB(byMeter,byLo,bySched,fileSum,patra,ptm)
        return summary, thruput
